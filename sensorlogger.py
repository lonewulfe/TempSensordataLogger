# -*- coding: utf-8 -*-
#
#   Author : Joel Louis
#
#

import threading, time, signal

from datetime import timedelta
import interelmodbus
import minimalmodbus
import serial
from openpyxl import Workbook
client1 = minimalmodbus.Instrument(port = "COM4", slaveaddress=1, mode='rtu', close_port_after_each_call=True, debug=False)
client2 = minimalmodbus.Instrument(port = "COM4", slaveaddress=2, mode='rtu', close_port_after_each_call=True, debug=False)
relay = minimalmodbus.Instrument(port = "COM4", slaveaddress=100, mode='rtu', close_port_after_each_call=True, debug=False)
passthrough = interelmodbus.Instrument(port = "COM4", slaveaddress=255, mode='rtu', close_port_after_each_call=True, debug=False)

#Details for Client1 (Temperature sensors, UI LEDs and OLED of Test EOS)
client1.serial.baudrate = 115200
client1.serial.timeout = 1
client1.serial.bytesize = 8
client1.serial.parity   = serial.PARITY_NONE
client1.serial.stopbits = 1
client1.serial.timeout  = 0.5
client1.mode = minimalmodbus.MODE_RTU
#Details for Client2 (Temperature Sensor of Reference EOS)
client2.serial.baudrate = 115200
client2.serial.timeout = 1
client2.serial.bytesize = 8
client2.serial.parity   = serial.PARITY_NONE
client2.serial.stopbits = 1
client2.serial.timeout  = 0.5
client2.mode = minimalmodbus.MODE_RTU
#Details for relay (Relays of EOS have the same slave address, make sure you disconnect the relay sections of the Reference EOS to avoid packet clash)
relay.serial.baudrate = 115200
relay.serial.timeout = 1
relay.serial.bytesize = 8
relay.serial.parity   = serial.PARITY_NONE
relay.serial.stopbits = 1
relay.serial.timeout  = 2.05
relay.mode = minimalmodbus.MODE_RTU
#Details for passthrough (This will use the modified library because minimalmodbus has an upper register limit of 127 while we need 255 and disabled CRC checks)
passthrough.serial.baudrate = 115200
passthrough.serial.timeout = 1
passthrough.serial.bytesize = 8
passthrough.serial.parity   = serial.PARITY_NONE
passthrough.serial.stopbits = 1
passthrough.serial.timeout  = 0.5
passthrough.mode = interelmodbus.MODE_RTU



WAIT_TIME_SECONDS = 30
t = 0
y = 2
#change this variable to change the duration of the test cases and cooldown intervals
minute = 60
hour = 60*60 #This is in seconds btw
#change the file path to actually save the file. Otherwise no one will be saved

class ProgramKilled(Exception):
    pass

def fun():
    sheet.cell(row=y,column=1).value = t

def signal_handler(signum, frame):
    raise ProgramKilled

class Job(threading.Thread):
    def __init__(self, interval, execute, *args, **kwargs):
        threading.Thread.__init__(self)
        self.daemon = False
        self.stopped = threading.Event()
        self.interval = interval
        self.execute = execute
        self.args = args
        self.kwargs = kwargs

    def stop(self):
                self.stopped.set()
                self.join()
    def run(self):
            while not self.stopped.wait(self.interval.total_seconds()):
                self.execute(*self.args, **self.kwargs)

workbook = Workbook()
sheet = workbook.active

sheet['A1'] = 'Time Elapsed (In Seconds)'
sheet['B1'] = 'Internal Temperature (In Celsius)'
sheet['C1'] = 'External Temperature (In Celsius)'
sheet['D1'] = 'Ambient Temperature (In Celsius)'
#Condition 0 activated
#passthrough._perform_command(249, '\x01\x00\x00\x03\x02\x4f\x46\x46')
#passthrough._perform_command(249, '\x01\x00\x00\x03\x01\x4f\x46\x46')
#passthrough._perform_command(249, '\x01\x00\x00\x03\x04\x4f\x46\x46')

print("Connected, Standby:")

if __name__ == "__main__":
    signal.signal(signal.SIGTERM, signal_handler)
    signal.signal(signal.SIGINT, signal_handler)
    job = Job(interval=timedelta(seconds=WAIT_TIME_SECONDS), execute=fun)
    job.start()

    while True:
        try:
            result1 = client1.read_registers(registeraddress = 16, number_of_registers = 2, functioncode=3) #Reading the internal temperature sensor of the test EOS
            result2 = client1.read_registers(registeraddress = 18, number_of_registers = 2, functioncode=3) #Reading the external temperature sensor of the test EOS
            result3 = client2.read_registers(registeraddress = 18, number_of_registers = 2, functioncode=3) #Reading the external temperature sensr of the ref EOS
            temp1 = result1[1]*0.001
            temp2 = result2[1]*0.001
            temp3 = result3[1]*0.001
            sheet.cell(row=y,column=2).value = temp1
            sheet.cell(row=y,column=3).value = temp2
            sheet.cell(row=y,column=4).value = temp3
            time.sleep(WAIT_TIME_SECONDS) #The actual delay function
            print(t,temp1,temp2,temp3)
            y+=1
            t+=30 #Thirty second interval count in Excel
#            if t == (120*minute):
#                client1.write_registers(registeraddress = 144, values = [3841])
#                sheet.cell(row=y,column=1).value = "OLED Test Active"
#                print("OLED Test Active")
#                y+=1
#            elif t == (121*minute):
#                client1.write_registers(registeraddress = 144, values = [0])
#                sheet.cell(row=y,column=1).value = "OLED Test Inactive"
#                print("OLED Test Inactive")
#                y+=1
#            elif t == (151*minute):
#                client1.write_registers(registeraddress = 80, values = [65535,65535])
#                client1.write_registers(registeraddress = 82, values = [65535,65535])
#                client1.write_registers(registeraddress = 84, values = [65535,65535])
#                client1.write_registers(registeraddress = 86, values = [65535,65535])
#                client1.write_registers(registeraddress = 88, values = [65535,65535])
#                client1.write_registers(registeraddress = 94, values = [65535,65535])
#                client1.write_registers(registeraddress = 96, values = [65535,65535])
#                client1.write_registers(registeraddress = 98, values = [65535,65535])
#                client1.write_registers(registeraddress = 100, values = [65535,65535])
#                client1.write_registers(registeraddress = 102, values = [65535,65535])
#                client1.write_registers(registeraddress = 104, values = [65535,65535])
#                client1.write_registers(registeraddress = 114, values = [65535,65535])
#                client1.write_registers(registeraddress = 116, values = [65535,65535])
#                client1.write_registers(registeraddress = 118, values = [65535,65535])
#                sheet.cell(row=y,column=1).value = "UI Test Active"
#                print("UI Test Active")
#                y+=1
#            elif t == (152*minute):
#                client1.write_registers(registeraddress = 80, values = [0,0])
#                client1.write_registers(registeraddress = 82, values = [0,0])
#                client1.write_registers(registeraddress = 84, values = [0,0])
#                client1.write_registers(registeraddress = 86, values = [0,0])
#                client1.write_registers(registeraddress = 88, values = [0,0])
#                client1.write_registers(registeraddress = 94, values = [0,0])
#                client1.write_registers(registeraddress = 96, values = [0,0])
#                client1.write_registers(registeraddress = 98, values = [0,0])
#                client1.write_registers(registeraddress = 100, values = [0,0])
#                client1.write_registers(registeraddress = 102, values = [0,0])
#                client1.write_registers(registeraddress = 104, values = [0,0])
#                client1.write_registers(registeraddress = 114, values = [0,0])
#                client1.write_registers(registeraddress = 116, values = [0,0])
#                client1.write_registers(registeraddress = 118, values = [0,0])
#                sheet.cell(row=y,column=1).value = "UI Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (182*minute):
#                client1.write_registers(registeraddress = 144, values = [3841])
#                client1.write_registers(registeraddress = 80, values = [65535,65535])
#                client1.write_registers(registeraddress = 82, values = [65535,65535])
#                client1.write_registers(registeraddress = 84, values = [65535,65535])
#                client1.write_registers(registeraddress = 86, values = [65535,65535])
#                client1.write_registers(registeraddress = 88, values = [65535,65535])
#                client1.write_registers(registeraddress = 94, values = [65535,65535])
#                client1.write_registers(registeraddress = 96, values = [65535,65535])
#                client1.write_registers(registeraddress = 98, values = [65535,65535])
#                client1.write_registers(registeraddress = 100, values = [65535,65535])
#                client1.write_registers(registeraddress = 102, values = [65535,65535])
#                client1.write_registers(registeraddress = 104, values = [65535,65535])
#                client1.write_registers(registeraddress = 114, values = [65535,65535])
#                client1.write_registers(registeraddress = 116, values = [65535,65535])
#                client1.write_registers(registeraddress = 118, values = [65535,65535])
#                sheet.cell(row=y,column=1).value = "UI & OLED Test Active"
#                print("UI & OLED Test Active")
#                y+=1
#            elif t == (183*minute):
#                client1.write_registers(registeraddress = 144, values = [0])
#                client1.write_registers(registeraddress = 80, values = [0,0])
#                client1.write_registers(registeraddress = 82, values = [0,0])
#                client1.write_registers(registeraddress = 84, values = [0,0])
#                client1.write_registers(registeraddress = 86, values = [0,0])
#                client1.write_registers(registeraddress = 88, values = [0,0])
#                client1.write_registers(registeraddress = 94, values = [0,0])
#                client1.write_registers(registeraddress = 96, values = [0,0])
#                client1.write_registers(registeraddress = 98, values = [0,0])
#                client1.write_registers(registeraddress = 100, values = [0,0])
#                client1.write_registers(registeraddress = 102, values = [0,0])
#                client1.write_registers(registeraddress = 104, values = [0,0])
#                client1.write_registers(registeraddress = 114, values = [0,0])
#                client1.write_registers(registeraddress = 116, values = [0,0])
#                client1.write_registers(registeraddress = 118, values = [0,0])
#                sheet.cell(row=y,column=1).value = "UI & OLED Test Inactive"
#                print("Cooldown Active")
#                y+=1
            if t == (4*hour):
                relay.write_register(registeraddress = 4, value = 1, number_of_decimals=0, functioncode=6, signed=False)
                sheet.cell(row=y,column=1).value = "Relay 1 Test Active"
                print("Relay 1 Test Active")
                y+=1
            elif t == (6*hour):
                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
                sheet.cell(row=y,column=1).value = "Relay 1 Test Inactive"
                print("Cooldown Active")
                y+=1


#            elif t == (21*minute):
#                relay.write_register(registeraddress = 4, value = 2, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 2 Test Active"
#                print("Relay 2 Test Active")
#                y+=1
#            elif t == (26*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 2 Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (31*minute):
#                relay.write_register(registeraddress = 4, value = 4, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 3 Test Active"
#                print("Relay 3 Test Active")
#                y+=1
#            elif t == (36*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 3 Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (41*minute):
#                relay.write_register(registeraddress = 4, value = 8, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 4 Test Active"
#                print("Relay 4 Test Active")
#                y+=1
#            elif t == (46*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 4 Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (51*minute):
#                relay.write_register(registeraddress = 4, value = 16, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 5 Test Active"
#                print("Relay 5 Test Active")
#                y+=1
#            elif t == (56*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 5 Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (61*minute):
#                relay.write_register(registeraddress = 4, value = 32, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 6 Test Active"
#                print("Relay 6 Test Active")
#                y+=1
#            elif t == (66*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "Relay 6 Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (453*minute):
#                relay.write_register(registeraddress = 4, value = 63, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "All Relays Test Active"
#                print("All Relays Test Active")
#                y+=1
#            elif t == (573*minute):
#                relay.write_register(registeraddress = 4, value = 0, number_of_decimals=0, functioncode=6, signed=False)
#                sheet.cell(row=y,column=1).value = "All Relays Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (81*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x02\x52\x45\x53\x45\x54')#ble_on
#                sheet.cell(row=y,column=1).value = "BLE Test Active"
#                print("BLE Test Active")
#                y+=1
#            elif t == (86*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x02\x4f\x46\x46') #ble_off
#                sheet.cell(row=y,column=1).value = "BLE Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (91*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x01\x52\x45\x53\x45\x54') #wifi_on
#                sheet.cell(row=y,column=1).value = "Wifi Test Active"
#                print("Wifi Test Active")
#                y+=1
#            elif t == (96*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x01\x4f\x46\x46') #wifi_off
#                sheet.cell(row=y,column=1).value = "Wifi Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (101*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x04\x52\x45\x53\x45\x54') #zigbee_on
#                sheet.cell(row=y,column=1).value = "Zigbee Test Active"
#                print("Zigbee Test Active")
#                y+=1
#            elif t == (106*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x04\x4f\x46\x46') #zigbee_off
#                sheet.cell(row=y,column=1).value = "Zigbee Test Inactive"
#                print("Cooldown Active")
#                y+=1
#            elif t == (693*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x01\x52\x45\x53\x45\x54')#wifi_on
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x02\x52\x45\x53\x45\x54')#ble_on
#                passthrough._perform_command(249, '\x01\x00\x00\x05\x04\x52\x45\x53\x45\x54')#zigbee_on
#                sheet.cell(row=y,column=1).value = "All Radio Chips Test Active"
#                print("All Radio Chips Test Active")
#                y+=1
#            elif t == (813*minute):
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x02\x4f\x46\x46')
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x01\x4f\x46\x46')
#                passthrough._perform_command(249, '\x01\x00\x00\x03\x04\x4f\x46\x46') #allradio_off
#                sheet.cell(row=y,column=1).value = "All Radio Chipset Test Inactive"
#                print("Cooldown Active")
#                y+=1
            elif t ==(10*hour):
                print("Time limit has been reached. Program has been terminated: Running CleanUp Code")
                job.stop()
                print("The file has been saved as EOSTempData6")
                filepath = "C:/Users/jl/OneDrive - Interel/Desktop/Test/EOSTempData6.xlsx"
                workbook.save(filepath)
                break

        except ProgramKilled:
            print ("Program killed: running cleanup code")
            job.stop()
            print("Name of the File:")
            x = input()
            filepath = "C:/Users/jl/OneDrive - Interel/Desktop/Test/" + x + ".xlsx"
            workbook.save(filepath)
            break
