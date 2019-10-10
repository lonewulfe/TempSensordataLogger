from openpyxl import Workbook
#name of workook = Workbook()
workbook = Workbook()
#This part is naming the active worksheet being used
sheet = workbook.active
#Define the titles here
sheet['A1'] = 'Time'
sheet['B1'] = 'Internal Temperature'
sheet['C1'] = 'External Temperature'
#Now we testing the timer thing, I feel like a hacker y'all this is so cool
y = 2
while (y < 6):
    sheet.cell(row=y,column=1).value = y-1
    y+=1
    pass
#define file paths before maybe? Just copy paste the address here with a filename in the end.
filepath="C:/Users/jl/OneDrive - Interel/Desktop/Test/demo.xlsx"
workbook.save(filepath)
